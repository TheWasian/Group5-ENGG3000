"""Collect ultrasonic calibration samples from the ESP32 over serial.

Install dependencies with:
    python3 -m pip install pyserial openpyxl matplotlib

Example:
    python3 calibration.py --port /dev/cu.usbmodemXXXX --count 10
"""

import argparse
import math
import time
from pathlib import Path

import matplotlib.pyplot as plt
import serial
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def collect_batch(connection: serial.Serial, count: int, timeout: float) -> tuple[list[float], list[float]]:
    """Request one batch and return sensor A and sensor B distances in cm."""
    connection.reset_input_buffer()
    connection.write(f"START,{count}\n".encode("ascii"))
    connection.flush()

    sensor_a: list[float] = []
    sensor_b: list[float] = []
    deadline = time.monotonic() + max(timeout, count * 0.2 + 5.0)
    began = False

    while time.monotonic() < deadline:
        raw_line = connection.readline()
        if not raw_line:
            continue

        line = raw_line.decode("ascii", errors="replace").strip()
        if line.startswith("BEGIN,"):
            began = True
            continue
        if line.startswith("ERROR,"):
            raise RuntimeError(line)
        if line == "DONE":
            if not began or len(sensor_a) != count or len(sensor_b) != count:
                raise RuntimeError(
                    f"Incomplete batch: received {len(sensor_a)} of {count} measurements"
                )
            return sensor_a, sensor_b
        if not line.startswith("DATA,"):
            continue

        fields = line.split(",")
        if len(fields) != 4:
            continue
        try:
            index = int(fields[1])
            distance_a = float(fields[2])
            distance_b = float(fields[3])
        except ValueError:
            continue

        if index != len(sensor_a) + 1:
            raise RuntimeError(f"Unexpected measurement index in line: {line}")
        if not math.isfinite(distance_a) or not math.isfinite(distance_b):
            raise RuntimeError(f"Invalid distance in line: {line}")
        sensor_a.append(distance_a)
        sensor_b.append(distance_b)

    raise TimeoutError("Timed out waiting for the ESP32 calibration batch")


def write_sheet(workbook: Workbook, title: str, actual_values: list[float], measurements: list[list[float]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(["Actual distance (cm)"] + [f"Measurement {index}" for index in range(1, len(measurements[0]) + 1)])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for actual_cm, values in zip(actual_values, measurements):
        sheet.append([actual_cm] + values)
    sheet.freeze_panes = "B2"
    sheet.column_dimensions["A"].width = 22
    for column in range(2, len(measurements[0]) + 2):
        sheet.column_dimensions[get_column_letter(column)].width = 16


def save_plot(output_path: Path, actual_values: list[float], sensor_data: dict[str, list[list[float]]]) -> None:
    figure, axis = plt.subplots(figsize=(10, 6))
    for sensor_name, batches in sensor_data.items():
        for actual_cm, values in zip(actual_values, batches):
            axis.scatter([actual_cm] * len(values), values, alpha=0.65, label=sensor_name if actual_cm == actual_values[0] else None)
        means = [sum(values) / len(values) for values in batches]
        axis.plot(actual_values, means, marker="o", label=f"{sensor_name} mean")
    axis.set_title("Ultrasonic Sensor Calibration")
    axis.set_xlabel("Actual distance (cm)")
    axis.set_ylabel("Measured distance (cm)")
    axis.grid(True, alpha=0.3)
    axis.legend()
    figure.tight_layout()
    figure.savefig(output_path, dpi=150)
    plt.show()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Collect ultrasonic sensor calibration data")
    parser.add_argument("--port", help="ESP32 serial port, for example /dev/cu.usbmodemXXXX")
    parser.add_argument("--baud", type=int, default=9600)
    parser.add_argument("--count", type=int, default=10, help="Measurements per actual distance")
    parser.add_argument("--output", type=Path, default=Path("ultrasonic_calibration.xlsx"))
    parser.add_argument("--sensor", choices=("A", "B", "both"), default="A")
    parser.add_argument("--timeout", type=float, default=10.0)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.count < 1 or args.count > 1000:
        raise SystemExit("--count must be between 1 and 1000")
    port = args.port or input("ESP32 serial port: ").strip()
    if not port:
        raise SystemExit("A serial port is required")

    actual_values: list[float] = []
    sensor_a_batches: list[list[float]] = []
    sensor_b_batches: list[list[float]] = []

    with serial.Serial(port, args.baud, timeout=1) as connection:
        time.sleep(2.0)
        connection.reset_input_buffer()
        print(f"Connected. Collecting {args.count} measurements per distance.")
        print("Place the target at a known distance, then enter it. Press Enter on a blank line to finish.")
        while True:
            entered = input("Actual distance in cm: ").strip()
            if not entered:
                break
            try:
                actual_cm = float(entered)
            except ValueError:
                print("Enter a number, such as 50 or 125.5.")
                continue
            if not math.isfinite(actual_cm) or actual_cm <= 0:
                print("Distance must be a positive number.")
                continue

            print("Measuring...")
            distances_a, distances_b = collect_batch(connection, args.count, args.timeout)
            actual_values.append(actual_cm)
            sensor_a_batches.append(distances_a)
            sensor_b_batches.append(distances_b)
            print(f"Recorded {args.count} measurements at {actual_cm:g} cm.")

    if not actual_values:
        raise SystemExit("No calibration data collected")

    workbook = Workbook()
    default_sheet = workbook.active
    assert default_sheet is not None
    workbook.remove(default_sheet)
    selected_data: dict[str, list[list[float]]] = {}
    if args.sensor in ("A", "both"):
        write_sheet(workbook, "Sensor A", actual_values, sensor_a_batches)
        selected_data["Sensor A"] = sensor_a_batches
    if args.sensor in ("B", "both"):
        write_sheet(workbook, "Sensor B", actual_values, sensor_b_batches)
        selected_data["Sensor B"] = sensor_b_batches

    workbook.save(args.output)
    plot_path = args.output.with_name(f"{args.output.stem}_plot.png")
    save_plot(plot_path, actual_values, selected_data)
    print(f"Saved spreadsheet: {args.output}")
    print(f"Saved plot: {plot_path}")


if __name__ == "__main__":
    main()
